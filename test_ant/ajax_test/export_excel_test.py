#coding=utf-8
'''
Created on 2016年8月18日

@author: MQM
'''
from openpyxl import Workbook
from openpyxl import load_workbook
import json

wb = Workbook() #新建文档
ws = wb.active
#workSheet = wb.create_sheet("sheet2")   #新建一页，名字“sheet2” 
#workSheet.title = "new sheet2"    #更改sheet 名字

#ws3 = wb["New Title"]    获取sheet，两种方式
#ws4 = wb.get_sheet_by_name("New Title")
ws['A1'] = "id"
ws['B1'] = "名称"
ws['C1'] = "备案编号"
ws['D1'] = "备案日期"
ws['E1'] = "备案状态"
ws['F1'] = "企业名称"
ws['G1'] = "企业地址"
ws['H1'] = "卫生许可"
ws['I1'] = "说明"
ws['J1'] = "processid"
for value in range(1,600000):
    
    with open('F:\JTPData\detail\\'+str(value)+'_gcft.json' , 'r',encoding='utf-8',errors='ignore') as f:
        jsonData = json.load(f)
        name = jsonData["productname"]
        apply_sn = jsonData["apply_sn"]
        provinceConfirm = jsonData["provinceConfirm"] #备案日期
        state = jsonData["state"]
        enterprise_name = jsonData["scqyUnitinfo"]["enterprise_name"]
        enterprise_address = jsonData["scqyUnitinfo"]["enterprise_address"]
        enterprise_healthpermits = jsonData["scqyUnitinfo"]["enterprise_healthpermits"]
        remark  = jsonData["scqyUnitinfo"]["remark"]
        processid  = jsonData["processid"]
        
        ws['A'+str(value+1)] = str(value)
        ws['B'+str(value+1)] = name
        ws['C'+str(value+1)] = apply_sn
        ws['D'+str(value+1)] = provinceConfirm
        ws['E'+str(value+1)] = state
        ws['F'+str(value+1)] = enterprise_name
        ws['G'+str(value+1)] = enterprise_address
        ws['H'+str(value+1)] = enterprise_healthpermits
        ws['I'+str(value+1)] = remark
        ws['J'+str(value+1)] = processid
        tep =int(ord('J'))
        for cf in jsonData["pfList"]:
            tep = tep+1
            cowNum = ''
            if tep <=90:
                cowNum = chr(tep)
            if tep > 90:
                cowNum = 'A'+chr(tep-26)
            if tep >116:
                cowNum = 'B'+chr(tep-52)
            if tep>142:
                cowNum = 'C'+chr(tep-78)
            if tep>168:
                cowNum ='D'+chr(tep-104)
            try:
                ws[cowNum+str(value+1)] = cf["cname"]
            except:
                print("行数超出")
            
    if value % 1000 == 0:
        print("value:",value)
        wb.save("F:\JTPData\\test.xlsx")
#wb.save("F:\JTPData\\test.xlsx")